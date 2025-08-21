import streamlit as st
import pandas as pd
import numpy as np
import joblib
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from datetime import datetime

def create_formatted_excel(df_input, waktu_prediksi, judul_sheet="Hasil Prediksi"):
    waktu_cetak = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    judul = "HASIL PREDIKSI KELAYAKAN PENERIMA BANTUAN PROGRAM KELUARGA HARAPAN (PKH) KELURAHAN CIPAMOKOLAN"

    wb = Workbook()
    ws = wb.active
    ws.title = judul_sheet

    # Judul utama
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_input.columns))
    ws["A1"] = judul
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Waktu prediksi
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df_input.columns))
    ws["A2"] = f"Waktu Prediksi: {waktu_prediksi}"
    ws["A2"].font = Font(italic=True, size=10)

    # Waktu cetak
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(df_input.columns))
    ws["A3"] = f"Waktu Cetak: {waktu_cetak}"
    ws["A3"].font = Font(italic=True, size=10)

    # ============ Header Tabel ============
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Warna kuning
    header_font = Font(bold=True)
    header_row = 4

    for col_num, column_title in enumerate(df_input.columns, 1):
        cell = ws.cell(row=header_row, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ============ Data Tabel ============
    for row_num, row_data in enumerate(df_input.values, start=header_row + 1):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # ============ Tambahkan Border ============
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row,
                            min_col=1, max_col=len(df_input.columns)):
        for cell in row:
            cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# === Load Model dan Preprocessing ===
model = joblib.load("model.pkl")
scaler = joblib.load("scaler.pkl")
selected_features = joblib.load("selected_features.pkl")
numerical_features = joblib.load("numerical_features.pkl")
categorical_features = joblib.load("categorical_features.pkl")
ohe = joblib.load("ohe.pkl")

st.set_page_config(page_title="Aplikasi PKH", layout="wide")
# Bikin dua kolom: satu untuk logo, satu untuk title
col1, col2 = st.columns([1, 3])  # Sesuaikan rasio kolom sesuai kebutuhan
with col1:
    st.image("logo.png", width=800)  # Ubah ukuran sesuai kebutuhan
with col2:
    st.title("Aplikasi Prediksi Kelayakan Penerima Bantuan Sosial Program Keluarga Harapan (PKH) Kelurahan Cipamokolan")
st.markdown("""
### 📖 Tentang Program Keluarga Harapan (PKH)

Program Keluarga Harapan (PKH) merupakan salah satu bentuk bantuan tunai bersyarat yang dikenal didunia sebagai
strategi efektif dalam menanggulangi kemiskinan kronis. Pelaksanaan PKH didasari oleh ketentuan hukum, salah satunya tercantum dalam Peraturan Menteri Sosial Nomor 1 Tahun
2018 tentang Program Keluarga Harapan.

Tujuan utama PKH adalah meningkatkan kualitas hidup keluarga miskin melalui kemudahan akses terhadap layanan pendidikan,
layanan kesehatan, serta pelayanan kesejahteraan sosial.
            
Kriteria penerima PKH meliputi:
- Ibu hamil atau menyusui
- Balita
- Anak prasekolah
- Anak sekolah
- Lanjut usia
- Penyandang disabilitas

Aplikasi ini membantu memprediksi apakah seseorang **layak atau tidak layak** menerima bantuan PKH berdasarkan data-data tersebut.
""")

# =====================
# 🔘 Mode Input
# =====================
mode = st.radio("Pilih Metode Prediksi", ["📝 Input Manual", "📁 Upload Excel"])

# ===========================
# 1️⃣ MODE INPUT MANUAL
# ===========================
if mode == "📝 Input Manual":
    st.subheader("🧾 Form Input Manual")

    input_data = {}

    # Numerik
    st.markdown("### 📌 Data Numerik")
    cols_num = st.columns(3)
    for i, col in enumerate(numerical_features):
        with cols_num[i % 3]:
            input_data[col] = st.number_input(f"{col}", step=1, format="%d")

    # Kategorikal
    st.markdown("### 📌 Data Kategorikal")
    cols_cat = st.columns(3)
    for i, col in enumerate(categorical_features):
        with cols_cat[i % 3]:
            if col.lower() == 'pekerjaan':
                options = ["PEDAGANG", "KARYAWAN SWASTA", "TIDAK ADA", "ASISTEN RUMAH TANGGA",
                        "BERTANI", "BURUH", "PEGAWAI", "PETERNAK", "WIRASWASTA"]
            elif col.lower() == 'status perkawinan':
                options = ["BELUM MENIKAH", "CERAI HIDUP", "CERAI MATI", "MENIKAH"]
            elif col.lower() == 'pendidikan':
                options = ["SD", "SMP", "SMA", "S1"]
            elif col.lower() == 'status rumah':
                options = ["BEBAS SEWA", "KONTRAK", "MENUMPANG", "MILIK ORANG TUA", "MILIK SENDIRI"]
            elif col.lower() == 'l/p':
                options = ["L", "P"]
            else:
                options = ["Ya", "Tidak"]
            input_data[col] = st.selectbox(f"{col}", options)

    if st.button("🔮 Prediksi"):
        input_df = pd.DataFrame([input_data])

        input_num = input_df[numerical_features]
        input_cat = input_df[categorical_features]
        encoded_cat = ohe.transform(input_cat)
        input_combined = np.hstack([input_num.values, encoded_cat])

        all_cols = numerical_features + list(ohe.get_feature_names_out(categorical_features))
        df_final = pd.DataFrame(input_combined, columns=all_cols)

        X_selected = df_final[selected_features]
        X_scaled = scaler.transform(X_selected)

        # 🔍 Prediksi dan Probabilitas
        pred = model.predict(X_scaled)
        proba = model.predict_proba(X_scaled)[0]
        label = "LAYAK" if pred[0] == 0 else "TIDAK LAYAK"
        prob_layak = round(proba[0] * 100, 2)
        prob_tidak_layak = round(proba[1] * 100, 2)

        st.success(f"✅ Hasil Prediksi: **{label}**")
        st.info(f"📊 Probabilitas: LAYAK = {prob_layak}%, TIDAK LAYAK = {prob_tidak_layak}%")

# ===========================
# 2️⃣ MODE UPLOAD EXCEL
# ===========================
elif mode == "📁 Upload Excel":
    st.subheader("📁 Unggah File Excel")
    uploaded_file = st.file_uploader("Upload file Excel (.xlsx)", type=["xlsx"])

    if uploaded_file:
        try:
            df_input = pd.read_excel(uploaded_file)
            st.write("📌 Kolom Terdeteksi:", df_input.columns.tolist())

            st.write("📄 **Data yang Diupload:**")
            st.dataframe(df_input)

            # Validasi kolom
            missing_cols = set(numerical_features + categorical_features) - set(df_input.columns)
            if missing_cols:
                st.warning(f"❗ Kolom berikut tidak ditemukan di file Excel: {', '.join(missing_cols)}")
            else:
                num_data = df_input[numerical_features]
                cat_data = df_input[categorical_features]

                encoded_cat = ohe.transform(cat_data)
                input_combined = np.hstack([num_data.values, encoded_cat])
                all_cols = numerical_features + list(ohe.get_feature_names_out(categorical_features))
                df_all = pd.DataFrame(input_combined, columns=all_cols)

                X_sel = df_all[selected_features]
                X_scaled = scaler.transform(X_sel)

                # 🔍 Prediksi dan Probabilitas
                pred = model.predict(X_scaled)
                proba_all = model.predict_proba(X_scaled)

                df_input['Hasil Prediksi'] = ["LAYAK" if p == 0 else "TIDAK LAYAK" for p in pred]
                df_input['Prob_LAYAK (%)'] = np.round(proba_all[:, 0] * 100, 2)
                df_input['Prob_TIDAK_LAYAK (%)'] = np.round(proba_all[:, 1] * 100, 2)
                waktu_prediksi = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

                st.success("✅ Prediksi selesai!")
                st.dataframe(df_input)

                # 🔍 Evaluasi Akurasi jika kolom label aktual tersedia
                if 'Status Kelayakan' in df_input.columns:
                    actual = df_input['Status Kelayakan'].str.upper().str.strip()
                    predicted = df_input['Hasil Prediksi'].str.upper().str.strip()

                    correct = (actual == predicted).sum()
                    total = len(df_input)
                    acc = correct / total * 100

                    st.markdown("### ✅ Evaluasi Akurasi")
                    st.info(f"Akurasi Prediksi: **{acc:.2f}%** ({correct} dari {total} prediksi benar)")
                else:
                    st.warning("⚠️ Kolom 'Status Kelayakan' tidak ditemukan, tidak bisa menghitung akurasi.")

                # ✅ Convert to formatted Excel in memory
                output = create_formatted_excel(df_input, waktu_prediksi)

                # ✅ Download Button
                st.download_button(
                    label="⬇️ Unduh Hasil sebagai Excel",
                    data=output,
                    file_name="hasil_prediksi_dengan_probabilitas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                warna_kategori = {
                     "LAYAK": "green",
                     "TIDAK LAYAK": "orange"
                }

                # 🔢 Visualisasi hasil prediksi dengan Plotly

                st.markdown("### 📊 Visualisasi Hasil Prediksi (Bar Chart)")
                hasil_counts = df_input['Hasil Prediksi'].value_counts().rename_axis('Kategori').reset_index(name='Jumlah')
                fig = px.bar(
                     hasil_counts,
                     x='Kategori',
                     y='Jumlah',
                     color='Kategori',
                     text='Jumlah',
                     title='Distribusi Hasil Prediksi',
                     color_discrete_map=warna_kategori
                     )
                fig.update_traces(textposition='outside')
                fig.update_layout(yaxis_title='Jumlah Data', xaxis_title='Kategori Prediksi')
                st.plotly_chart(fig, use_container_width=True)

                # 📊 Visualisasi Hasil Prediksi: Pie Chart
                
                st.markdown("### 🥧 Visualisasi Hasil Prediksi (Pie Chart)")
                fig = px.pie(
                      hasil_counts,
                      names='Kategori',
                      values='Jumlah',
                      title='Distribusi Prediksi Kelayakan',
                      color='Kategori',
                      color_discrete_map=warna_kategori
                      )
                fig.update_traces(textposition='inside', textinfo='percent+label')
                st.plotly_chart(fig, use_container_width=True)

                if 'Penghasilan' in df_input.columns:
                    fig = px.box(
                        df_input,
                        x='Hasil Prediksi',
                        y='Penghasilan',
                        color='Hasil Prediksi',
                        title='Distribusi Penghasilan Berdasarkan Hasil Prediksi',
                        color_discrete_map=warna_kategori
                    )
                    st.plotly_chart(fig, use_container_width=True)

                    if 'Pendidikan' in df_input.columns:
                        pend_counts = df_input.groupby(['Pendidikan', 'Hasil Prediksi']).size().reset_index(name='Jumlah')
                        fig = px.bar(
                        pend_counts,
                        x='Pendidikan',
                        y='Jumlah',
                        color='Hasil Prediksi',
                        barmode='group',
                        title='Pendidikan vs Hasil Prediksi',
                        color_discrete_map=warna_kategori
                    )
                    st.plotly_chart(fig, use_container_width=True)

                    if 'Status Rumah' in df_input.columns:
                        rumah_counts = df_input.groupby(['Status Rumah', 'Hasil Prediksi']).size().reset_index(name='Jumlah')
                        fig = px.bar(
                        rumah_counts,
                        x='Status Rumah',
                        y='Jumlah',
                        color='Hasil Prediksi',
                        barmode='group',
                        title='Status Rumah vs Hasil Prediksi',
                        color_discrete_map=warna_kategori
                    )
                    st.plotly_chart(fig, use_container_width=True)

        except Exception as e:
            st.error(f"❌ Terjadi kesalahan saat memproses file: {e}")